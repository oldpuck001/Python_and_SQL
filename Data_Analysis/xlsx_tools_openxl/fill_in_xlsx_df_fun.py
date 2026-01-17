# fill_in_xlsx_df_fun.py

# 从指定位置开始填写整个工作表

import openpyxl

def fill_in_xlsx_df(file_path, fill_sheet, fill_df, start_row=1, start_col=1, data_only=False):

    wb =  openpyxl.load_workbook(file_path, data_only=data_only)
    ws = wb[fill_sheet]

    for j, col in enumerate(fill_df.columns):
        ws.cell(row=start_row, column=start_col + j, value=col)
    data_start = start_row + 1

    for i, row in enumerate(fill_df.values):
        for j, value in enumerate(row):
            ws.cell(row=data_start+i, column=start_col+j, value=value)

    wb.save(file_path)

    return 'The xlsx output file was updata successfully!\n'