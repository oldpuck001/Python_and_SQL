# export_tree_new_xlsx_fun.py

# 导出新树状图 xlsx 文件

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from xlsx_tools_openxl import cell_format_fun

def export_tree_new_xlsx(target_file_path,
                         in_groupby_summary_df, in_groupby_summary_df_label, in_priority_list, in_group_by_summary_dict,
                         in_summary_total, out_summary_total,
                         out_groupby_summary_df, out_groupby_summary_df_label, out_priority_list, out_group_by_summary_dict
                         ):

    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "树状图"

    # 添加In数据到左边
    in_cycle_not_sorted = list(in_groupby_summary_df[in_groupby_summary_df_label])
    ordered = [item for item in in_priority_list if item in in_cycle_not_sorted]
    remaining = [item for item in in_cycle_not_sorted if item not in ordered]
    in_cycle = ordered + remaining

    in_step = 1
    in_skip = 0

    for in_groupby in in_cycle:
        in_step += in_skip
        col_c = ws.cell(row=in_step, column=3, value=in_groupby)
        cell_format_fun.cell_format(col_c, 2)
        in_groupby_value = in_groupby_summary_df.query('in_groupby_sort == @in_groupby')['in_groupby_value'].values[0]
        col_d = ws.cell(row=in_step, column=4, value=in_groupby_value)
        cell_format_fun.cell_format(col_d, 2)

        in_skip = len(in_group_by_summary_dict[in_groupby]) + 1

        secondary_in_step = in_step
        for row in dataframe_to_rows(in_group_by_summary_dict[in_groupby], index=False, header=False):

            col_a = ws.cell(row=secondary_in_step, column=1, value=row[0])
            cell_format_fun.cell_format(col_a, 2)
            col_b = ws.cell(row=secondary_in_step, column=2, value=row[1])
            cell_format_fun.cell_format(col_b, 2)
            secondary_in_step += 1

    # 中间部分
    col_E1 = ws.cell(row=1, column=5, value='In值合计')
    cell_format_fun.cell_format(col_E1, 2)
    col_E2 = ws.cell(row=2, column=5, value=in_summary_total)
    cell_format_fun.cell_format(col_E2, 2)

    col_F1 = ws.cell(row=1, column=6, value='Out值合计')
    cell_format_fun.cell_format(col_F1, 2)
    col_F2 = ws.cell(row=2, column=6, value=out_summary_total)
    cell_format_fun.cell_format(col_F2, 2)

    # 添加Out数据到右边
    out_cycle_not_sorted = list(out_groupby_summary_df[out_groupby_summary_df_label])
    ordered = [item for item in out_priority_list if item in out_cycle_not_sorted]
    remaining = [item for item in out_cycle_not_sorted if item not in ordered]
    out_cycle = ordered + remaining

    out_step = 1
    out_skip = 0

    for out_groupby in out_cycle:
        out_step += out_skip
        col_g = ws.cell(row=out_step, column=7, value=out_groupby)
        cell_format_fun.cell_format(col_g, 2)
        out_groupby_value = out_groupby_summary_df.query('out_groupby_sort == @out_groupby')['out_groupby_value'].values[0]
        col_h = ws.cell(row=out_step, column=8, value=out_groupby_value)
        cell_format_fun.cell_format(col_h, 2)

        out_skip = len(out_group_by_summary_dict[out_groupby]) + 1
        
        secondary_out_step = out_step
        for row in dataframe_to_rows(out_group_by_summary_dict[out_groupby], index=False, header=False):

            col_i = ws.cell(row=secondary_out_step, column=9, value=row[0])
            cell_format_fun.cell_format(col_i, 2)
            col_j = ws.cell(row=secondary_out_step, column=10, value=row[1])
            cell_format_fun.cell_format(col_j, 2)
            secondary_out_step += 1

    # 保存 Excel 文件
    wb.save(target_file_path)

    return '导出成功！\n'