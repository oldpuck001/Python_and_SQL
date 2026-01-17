# fill_in_xlsx_list_fun.py

# 逐个填写工作表单元格

import openpyxl
from openpyxl.utils import get_column_letter
from xlsx_tools_openxl import cell_format_fun

def fill_in_xlsx_list(file_path, fill_in_xlsx, data_only=False):

    wb = openpyxl.load_workbook(file_path, data_only=data_only)
    ws = wb[fill_in_xlsx[0][0]]

    # 起始单元格位置
    r = fill_in_xlsx[0][1]
    c = fill_in_xlsx[0][2]

    # 设置行高
    r_n = r + 1
    for n in fill_in_xlsx[1]:
        ws.row_dimensions[r_n].height = n
        r_n += 1

    # 设置列宽
    c_n = c
    for n in fill_in_xlsx[2]:
        ws.column_dimensions[get_column_letter(c_n)].width = n
        c_n += 1

    # 填写单元格、设置单元格格式
    r_f = r
    for sheet_row in fill_in_xlsx[3:]:
        r_f += 1
        c_f = c
        for sheet_cell in sheet_row:
            cell_fill = ws.cell(row=r_f, column=c_f, value=sheet_cell[0])
            cell_format_fun.cell_format(cell_fill, sheet_cell[1])
            c_f += 1

    wb.save(file_path)

    return 'The xlsx output file was updata successfully!\n'