# OpenPyXL_excel.py

import pandas as pd
import openpyxl
import openpyxl.styles
import datetime as dt

file_path = '/Users/lei/Downloads/stores.xlsx'

# 使用OpenPyXL读取文件
# 使用openpyxl.load_workbook打开工作簿来读取单元格的值時，在加载数据之后文件会自动关闭
# data_only的默认值是False，此时会返回单元格的公式
# 使用 data_only=True 時，获得的是单元格的值
book = openpyxl.load_workbook(file_path, data_only=True)

# 获取所有工作表名称的列表
print(book.sheetnames)

# 通过名称或索引（从0开始）获取工作表对象
sheet = book['Sheet1']
sheet = book.worksheets[0]

# 遍历所有工作表对象，openpyxl使用的是title而不是name
for i in book.worksheets:
    print(i.title)

# 获取维度，以工作表所选区域为例
print(sheet.max_row, sheet.max_column)

# 读取单个单元格的值，分别使用的是A1这种表示法，以及单元格索引（从1开始）
print(sheet['B6'].value)
print(sheet.cell(row=6, column=2).value)


# 使用OpenPyXL写入文件
# 实例化工作簿
book = openpyxl.Workbook()

# 获取第一张工作表并赋予它一个名称
sheet = book.active
sheet.title = 'Sheet1'

# 使用A1表示法和单元格索引,（从1开始）写入各个单元格
sheet['A1'].value = 'Hello, world!'
sheet.cell(row=2, column=1, value='Hello, my world!')

# 格式化
sheet['C4'].value = 'Hello, our world!'

# 字體顏色、加粗
sheet['C4'].font = openpyxl.styles.Font(color='FF0000', bold=True)

# 邊框
thin = openpyxl.styles.Side(border_style='thin')
sheet['C4'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

# 居中
sheet['C4'].alignment = openpyxl.styles.Alignment(horizontal="center")

# 填充：黃色、橘色、綠色
sheet['E4'].fill = openpyxl.styles.PatternFill(fgColor='FFFF00', fill_type='solid')
sheet['F4'].fill = openpyxl.styles.PatternFill(fgColor='FFC000', fill_type='solid')
sheet['G4'].fill = openpyxl.styles.PatternFill(fgColor='92D050', fill_type='solid')

# 数字格式化（使用Excel的格式化字符串）
sheet['A5'].value = 33333.3333
sheet['A5'].number_format = '#,##0.00'

# 数字格式化（使用Excel的格式化字符串）
sheet['A6'].value = -55555.36
sheet['A6'].number_format = '#,##0.00'


# 公式：必须使用以逗号分隔的英文公式名称
sheet['A7'].number_format = '#,##0.00'
sheet['A7'].value = '=SUM(A5:A6)'

# 日期格式化（使用Excel的格式化字符串）
sheet['A8'].value = dt.date(2016, 10, 13)
sheet['A8'].number_format = 'yyyy-mm-dd'

# 保存工作簿会在磁盘上创建文件
book.save('/Users/lei/Downloads/openpyxl.xlsx')


# 讀取後保存，等於編輯的效果
# 读取stores.xlsx文件，修改一个单元格，并将其以新的名称保存到新的位置
book = openpyxl.load_workbook('/Users/lei/Downloads/openpyxl.xlsx')
book['Sheet1']['A10'].value = 'modified'
book.save('/Users/lei/Downloads/openpyxl_edited.xlsx')