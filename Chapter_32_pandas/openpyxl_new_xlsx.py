# openpyxl_new_xlsx.py

import openpyxl
import openpyxl.styles
import datetime as dt


# 使用OpenPyXL写入文件
# 实例化工作簿
wb = openpyxl.Workbook()

# 获取第一张工作表并赋予它一个名称
ws = wb.active
ws.title = 'Sheet1'

# 设置标题
ws.merge_cells('A1:F1')
ws['A1'].font = openpyxl.styles.Font(size=24, bold=True)
ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

# 使用A1表示法和单元格索引,（从1开始）写入各个单元格
ws['A1'].value = 'Hello, world!'
ws.cell(row=2, column=1, value='Hello, my world!')

# 在循環中給單元格輸入值及設置格式
value_list = ['项目', '期末余额', '上年末余额', '项目', '期末余额', '上年末余额']
start_col = 1
for p in value_list:
    ws.cell(row=10, column=start_col, value=p)
    start_col += 1

for row in ws['A10:F10']:
    for cell in row:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')


# 格式化
ws['C4'].value = 'Hello, our world!'

# 字體顏色、加粗
ws['C4'].font = openpyxl.styles.Font(color='FF0000', bold=True)

# 邊框
thin = openpyxl.styles.Side(border_style='thin')
ws['C4'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

thin_border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                     right=openpyxl.styles.Side(style='thin'),
                                     top=openpyxl.styles.Side(style='thin'),
                                     bottom=openpyxl.styles.Side(style='thin'))

for row in ws.iter_rows(min_row=4, min_col=5, max_row=9, max_col=10):
    for cell in row:
        cell.border = thin_border

# 居中
ws['C4'].alignment = openpyxl.styles.Alignment(horizontal="center")

# 自动换行
ws['C4'].alignment = openpyxl.styles.Alignment(wrap_text=True)

# 设置列宽
ws.column_dimensions['A'].width = 28

# 填充：黃色、橘色、綠色
ws['E4'].fill = openpyxl.styles.PatternFill(fgColor='FFFF00', fill_type='solid')
ws['F4'].fill = openpyxl.styles.PatternFill(fgColor='FFC000', fill_type='solid')
ws['G4'].fill = openpyxl.styles.PatternFill(fgColor='92D050', fill_type='solid')

# 数字格式化（使用Excel的格式化字符串）
ws['A5'].value = 33333.3333
ws['A5'].number_format = '#,##0.00'

# 数字格式化（使用Excel的格式化字符串）
ws['A6'].value = -55555.36
ws['A6'].number_format = '#,##0.00'


# 公式：必须使用以逗号分隔的英文公式名称
ws['A7'].number_format = '#,##0.00'
ws['A7'].value = '=SUM(A5:A6)'

# 日期格式化（使用Excel的格式化字符串）
ws['A8'].value = dt.date(2016, 10, 13)
ws['A8'].number_format = 'yyyy-mm-dd'


# 保存工作簿会在磁盘上创建文件
wb.save('/Users/lei/Downloads/openpyxl.xlsx')




# 讀取後保存，等於編輯的效果
# 读取stores.xlsx文件，修改一个单元格，并将其以新的名称保存到新的位置
book = openpyxl.load_workbook('/Users/lei/Downloads/openpyxl.xlsx')
book['Sheet1']['A10'].value = 'modified'
book.save('/Users/lei/Downloads/openpyxl_edited.xlsx')