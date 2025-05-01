# pandas_and_OpenPyXL.py

import pandas as pd
import openpyxl.styles

# 讀取值
with pd.ExcelFile('stores.xlsx', engine='openpyxl') as xlfile:
    # 读取DataFrame
    df = pd.read_excel(xlfile, sheet_name='Sheet1')
    print(df)

    # 获取OpenPyXL工作簿对象
    book = xlfile.book
    # OpenPyXL代码从这里开始
    value = book['Sheet2']['B3'].value
    print(value)

# 寫入值
with pd.ExcelWriter('/Users/lei/Downloads/pandas_and_openpyxl.xlsx', engine='openpyxl') as writer:
    df = pd.DataFrame({'col1': [1, 2, 3, 4], 'col2': [5, 6, 7, 8]})
    # 写入DataFrame
    df.to_excel(writer, 'Sheet1', startrow=4, startcol=2)

    # 获取OpenPyXL工作簿和工作表对象
    book = writer.book
    sheet = writer.sheets['Sheet1']
    # OpenPyXL的代码从这里开始
    sheet['A1'].value = 'This is a Title'


# 設置格式
df = pd.DataFrame({'col1': [1, -2], 'col2': [-3, 4]}, index=['row1', 'row2'])
df.index.name = 'ix'

# 用OpenPyXL格式化索引和标题
with pd.ExcelWriter('/Users/lei/Downloads/formatting_openpyxl.xlsx', engine='openpyxl') as writer:
    # 将整个df以默认格式从A1处写入
    df.to_excel(writer, startrow=0, startcol=0)
    # 将整个df以默认自定义索引/标题格式从A6处写入
    startrow, startcol = 0, 5
    # 写入DataFrame的数据部分
    df.to_excel(writer, header=False, index=False, startrow=startrow + 1, startcol=startcol + 1)

    # 获取工作表对象并创建样式对象
    sheet = writer.sheets['Sheet1']
    style = openpyxl.styles.PatternFill(fgColor="D9D9D9", fill_type="solid")
    # 写入带样式的列标题
    for i, col in enumerate(df.columns):
        sheet.cell(row=startrow + 1, column=i + startcol + 2, value=col).fill = style
    # 写入带样式的索引
    index = [df.index.name if df.index.name else None] + list(df.index)
    for i, row in enumerate(index):
        sheet.cell(row=i + startrow + 1, column=startcol + 1, value=row).fill = style

# 格式化 DataFrame 的数据部分
with pd.ExcelWriter('/Users/lei/Downloads/data_format_openpyxl.xlsx', engine='openpyxl') as writer:
    # 写入DataFrame
    df.to_excel(writer)
    # 获取工作簿对象和工作表对象
    book = writer.book
    sheet = writer.sheets['Sheet1']
    # 格式化每一个单元格
    nrows, ncols = df.shape
    for row in range(nrows):
        for col in range(ncols):
            # 考虑到标题/索引，这里加1
            # 因为OpenPyXL的索引是从1开始的，所以还要加1
            cell = sheet.cell(row=row + 2, column=col + 2)
            cell.number_format = "0.000"
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")