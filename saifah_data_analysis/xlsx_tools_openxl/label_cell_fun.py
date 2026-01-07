# label_cell_fun.py

import openpyxl

# 标记工作表单元格
def label_cell(file_path, sheet_name, label_list):

    wb = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb[sheet_name]

    # 标记单元格（黄色）
    for label in label_list:
        ws.cell(row=label[0], column=label[1]).fill = openpyxl.styles.PatternFill(fgColor='FFFF00', fill_type='solid')

    wb.save(file_path)

    info = 'Successful.'

    return [True, info]