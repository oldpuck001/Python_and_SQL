# gui_tk_sheet_comparision.py

import subprocess
import pandas as pd
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from gui_tk import gui_tk_area_text
from dataframe_tools_pd import sheetnames_import_fun
from dataframe_tools_pd import read_xlsx_xls_csv_txt_fun
from xlsx_tools_openxl import export_new_xlsx_fun
from xlsx_tools_openxl import label_cell_fun

class gui_tk_sheet_comparision_class:

    def gui_tk_sheet_comparision_frame(self, root, control_frame_config, text_area):
        
        frame_result = tk.Frame(root)
        frame_result.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=5, pady=5)

        options = []

        frame_result.frame_1 = tk.Frame(frame_result)
        frame_result.frame_1.pack(side=tk.TOP, fill=tk.BOTH)
        tk.Label(frame_result.frame_1, text='File Path 1', width=10).pack(side=tk.LEFT, padx=5, pady=5)
        frame_result.frame_1.entry_file_1 = tk.Entry(frame_result.frame_1, state='readonly', readonlybackground='white')         # 创建Entry并保存引用
        frame_result.frame_1.entry_file_1.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5, pady=5)

        frame_result.frame_2 = tk.Frame(frame_result)
        frame_result.frame_2.pack(side=tk.TOP, fill=tk.BOTH)
        tk.Label(frame_result.frame_2, text='选择工作表', width=10).pack(side=tk.LEFT, padx=5, pady=5)
        frame_result.frame_2.combobox_sheet_1 = ttk.Combobox(frame_result.frame_2, values=options, state='readonly')
        frame_result.frame_2.combobox_sheet_1.pack(side=tk.LEFT, padx=5, pady=5)

        frame_result.frame_3 = tk.Frame(frame_result)
        frame_result.frame_3.pack(side=tk.TOP, fill=tk.BOTH)
        tk.Label(frame_result.frame_3, text='File Path 2', width=10).pack(side=tk.LEFT, padx=5, pady=5)
        frame_result.frame_3.entry_file_2 = tk.Entry(frame_result.frame_3, state='readonly', readonlybackground='white')         # 创建Entry并保存引用
        frame_result.frame_3.entry_file_2.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5, pady=5)

        frame_result.frame_4 = tk.Frame(frame_result)
        frame_result.frame_4.pack(side=tk.TOP, fill=tk.BOTH)
        tk.Label(frame_result.frame_4, text='选择工作表', width=10).pack(side=tk.LEFT, padx=5, pady=5)
        frame_result.frame_4.combobox_sheet_2 = ttk.Combobox(frame_result.frame_4, values=options, state='readonly')
        frame_result.frame_4.combobox_sheet_2.pack(side=tk.LEFT, padx=5, pady=5)

        frame_result.frame_5 = tk.Frame(frame_result)
        frame_result.frame_5.pack(side=tk.TOP, fill=tk.BOTH)
        tk.Button(frame_result.frame_5, text=control_frame_config['button_name'][0],
                  command=lambda: self.input_sheet(frame_result.frame_1.entry_file_1,
                                                   frame_result.frame_2.combobox_sheet_1,
                                                   text_area),
                  width=10).pack(side=tk.LEFT, padx=5)
        tk.Button(frame_result.frame_5, text=control_frame_config['button_name'][1],
                  command=lambda: self.input_sheet(frame_result.frame_3.entry_file_2,
                                                   frame_result.frame_4.combobox_sheet_2,
                                                   text_area),
                  width=10).pack(side=tk.LEFT, padx=5)
        tk.Button(frame_result.frame_5, text=control_frame_config['button_name'][2],
                  command=lambda: self.compare_sheet(frame_result.frame_1.entry_file_1,
                                                     frame_result.frame_2.combobox_sheet_1,
                                                     frame_result.frame_3.entry_file_2,
                                                     frame_result.frame_4.combobox_sheet_2,
                                                     text_area),
                  width=10).pack(side=tk.LEFT, padx=5)

        return frame_result
    

    # 导入按钮函数
    def input_sheet(self, entry_file, combobox_sheet, text_area):

        fill_text = ''
        path = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx'),
                                                     ('Excel Files', '*.xls')])

        if path:
            entry_file.config(state='normal')
            entry_file.delete(0, tk.END)
            entry_file.insert(0, path)
            entry_file.config(state='readonly')
            fill_text += f'Selected: {path}\n'
            fill_text += 'File selection successful!\n'

            sheets_name_result = sheetnames_import_fun.sheetnames_import(path)
            if sheets_name_result[0]:

                sheets_name_list = sheets_name_result[2]
                combobox_sheet['values'] = sheets_name_list
                combobox_sheet.set(sheets_name_list[0])               # 设置默认选择第一个
                combobox_sheet.config(state='readonly')

            else:

                combobox_sheet.set('')
                combobox_sheet.config(state='disabled')

            fill_text += sheets_name_result[1]

        else:
            fill_text += f'File selection failed!\n'

        gui_tk_area_text.text_area_fill(text_area, fill_text)


    # 开始对比按钮函数
    def compare_sheet(self, entry_file_1, combobox_sheet_1, entry_file_2, combobox_sheet_2, text_area):

        fill_text = ''
        file_path_1 = entry_file_1.get()
        sheet_name_1 = combobox_sheet_1.get()
        file_path_2 = entry_file_2.get()
        sheet_name_2 = combobox_sheet_2.get()

        # 读入数据
        result_info_1 = read_xlsx_xls_csv_txt_fun.read_xlsx_xls_csv_txt(file_path=file_path_1, sheet_name=sheet_name_1)
        result_info_2 = read_xlsx_xls_csv_txt_fun.read_xlsx_xls_csv_txt(file_path=file_path_2, sheet_name=sheet_name_2)
        if result_info_1[0]:
            df_1 = result_info_1[2]
        if result_info_2[0]:
            df_2 = result_info_2[2]

        # 對齊資料行列數
        max_rows = max(df_1.shape[0], df_2.shape[0])
        max_cols = max(df_1.shape[1], df_2.shape[1])

        diffs = []
        for row in range(max_rows):
            for col in range(max_cols):
                val1 = df_1.iat[row, col] if row < df_1.shape[0] and col < df_1.shape[1] else None
                val2 = df_2.iat[row, col] if row < df_2.shape[0] and col < df_2.shape[1] else None
                if (val1 != val2) and (not pd.isnull(val1) and not pd.isnull(val2)):
                    diffs.append({
                        'Row': row + 2,
                        'Column': col + 1,
                        'File1': val1,
                        'File2': val2
                    })

        df_result = pd.DataFrame(diffs)

        # 顯示差異
        if diffs:

            label_list = []
            for n in range(len(diffs)):
                label_list.append([df_result['Row'][n], df_result['Column'][n]])
            label_cell_fun.label_cell(file_path_1, sheet_name_1, label_list)
            label_cell_fun.label_cell(file_path_2, sheet_name_2, label_list)

            df_export = df_result.copy()
            df_export['Column'] = df_export['Column'].map(get_column_letter)

            result_info = export_new_xlsx_fun.export_new_xlsx(df_export)

            if result_info[0]:
                fill_text += result_info[1]
                subprocess.run(['open', result_info[2]])

            fill_text += f'原表格标记完成，差异汇总表导出成功，文件路径：{result_info[2]}'

        else:

            fill_text += '两个表格完全一致。'

        gui_tk_area_text.text_area_fill(text_area, fill_text)