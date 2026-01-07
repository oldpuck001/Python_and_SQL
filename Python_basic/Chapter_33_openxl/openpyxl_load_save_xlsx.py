# openpyxl_load_save_xlsx.py

import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter


file_path = '/Users/lei/Downloads/stores.xlsx'

# 使用OpenPyXL读取文件
# 使用openpyxl.load_workbook打开工作簿来读取单元格的值時，在加载数据之后文件会自动关闭
# data_only的默认值是False，此时会返回单元格的公式
# 使用 data_only=True 時，获得的是单元格的值
wb = openpyxl.load_workbook(file_path, data_only=True)

# 获取所有工作表名称的列表
print(wb.sheetnames)


# 默认激活的工作表是第一个，也就是 'Sheet1'
ws = wb.active

# 通过名称或索引（从0开始）获取工作表对象
sheet = wb['Sheet1']
sheet = wb.worksheets[0]

# 遍历所有工作表对象，openpyxl使用的是title而不是name
for i in wb.worksheets:
    print(i.title)

# 获取维度，以工作表所选区域为例
print(sheet.max_row, sheet.max_column)

# 读取单个单元格的值，分别使用的是A1这种表示法，以及单元格索引（从1开始）
print(sheet['B6'].value)
print(sheet.cell(row=6, column=2).value)


# 添加边框
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = thin_border

# 添加筛选功能，筛选行设置在第一行
ws.auto_filter.ref = ws.dimensions

# 设置日期和数值格式，遍历所有列
for col in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in col:
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'             # 数字格式设置为千分位和保留两位小数

# 设置凭证字号列居中
#for row in range(2, ws.max_row + 1):
#    ws.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='center')

# 调整列宽（按实际列顺序）
column_widths = {
    '一级科目编号': 15,
    '一级科目名称': 20,
    '二级科目编号': 15,
    '二级科目名称': 50,
    '期初借方': 17,
    '期初贷方': 17,
    '本期借方': 17,
    '本期贷方': 17,
    '期末借方': 17,
    '期末贷方': 17
}

# 设置列宽
for i, (col_name, width) in enumerate(column_widths.items(), start=1):
    col_letter = get_column_letter(i)
    ws.column_dimensions[col_letter].width = width

# 保存更改
wb.save(file_path)
wb.close()


def convert_to_numeric(value):
    try:
        return pd.to_numeric(value.replace(',', ''))
    except AttributeError:
        return 0 if pd.isna(value) else value


def determine_subject_length(code):
    return len(str(code).strip())